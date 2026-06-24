from __future__ import annotations

import csv
import re
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
FORBIDDEN_LAYOUT = re.compile(r"[\r\n]| {2,}")
DEPENDENT_WORDS = re.compile(r"[①②③④⑤⑥⑦⑧⑨⑩㈱㈲㍾㍽㍼㍻]" )
NUMERIC_WITH_DECORATION = re.compile(r"^\s*[+-]?[0-9][0-9, ]*(?:\.[0-9]+)?\s*(?:[%％円人件戸\*†])\s*$")
ERA_ONLY = re.compile(r"^(?:令和|平成|昭和|大正|明治)\s*\d+(?:年)?$")


@dataclass(frozen=True)
class Finding:
    code: str
    message: str
    severity: str = "warning"
    row: int | None = None
    column: int | None = None
    value: str | None = None


@dataclass
class CheckResult:
    path: str
    findings: list[Finding] = field(default_factory=list)

    @property
    def valid(self) -> bool:
        return not any(item.severity == "error" for item in self.findings)

    def as_dict(self) -> dict:
        checks = _check_statuses(self.path, self.findings)
        return {
            "path": self.path,
            "valid": self.valid,
            "findings": [_finding_as_dict(item) for item in self.findings],
            "checks": checks,
            "summary": {
                "issues_found": sum(item["status"] == "issues_found" for item in checks),
                "passed": sum(item["status"] == "passed" for item in checks),
                "not_applicable": sum(item["status"] == "not_applicable" for item in checks),
            },
        }


# 表記は e-Stat「結果表における機械判読可能なデータ作成に関する表記方法 Ver.1.2」の
# チェック項目名から抜粋している。
CHECK_ITEM_REFERENCES = {
    "missing-header": "チェック項目２-５ 項目名等を省略していないか",
    "duplicate-header": "チェック項目２-５ 項目名等を省略していないか",
    "inconsistent-columns": "チェック項目４-５ １行１データで表現されているか",
    "leading-empty-rows": "チェック項目４-１３ データが分断されていないか",
    "split-table": "チェック項目４-１３ データが分断されていないか",
    "layout-whitespace": "チェック項目２-４ スペースや改行等で体裁を整えていないか",
    "dependent-character": "チェック項目２-９ 機種依存文字を使用していないか。",
    "decorated-number": "チェック項目２-２ 数値データは数値属性とし、文字列を含まないこと",
    "era-only-date": "チェック項目２-10 西暦表記又は和暦に西暦の併記がされているか",
    "merged-cells": "チェック項目２-３ セルの結合をしていないか",
    "formulas": "チェック項目２-６ 数式を使用している場合は、数値データに修正しているか",
}


def _finding_as_dict(finding: Finding) -> dict:
    result = asdict(finding)
    result["check_item"] = CHECK_ITEM_REFERENCES.get(finding.code)
    return result


CHECK_GROUPS = (
    ("file-format", "ファイル形式・読み取り", ("unsupported-format", "invalid-xlsx", "empty-workbook", "legacy-xls"), "all"),
    ("headers", "項目名の欠落・重複", ("missing-header", "duplicate-header"), "table"),
    ("table-structure", "列数・空白行などの表構造", ("leading-empty-rows", "inconsistent-columns", "split-table", "empty-table"), "table"),
    ("layout", "空白・改行による体裁調整", ("layout-whitespace",), "table"),
    ("characters", "機種依存文字", ("dependent-character",), "table"),
    ("numbers", "数値と単位・記号の混在", ("decorated-number",), "table"),
    ("dates", "和暦のみの時間軸", ("era-only-date",), "table"),
    ("xlsx-merged-cells", "XLSX のセル結合", ("merged-cells",), "xlsx"),
    ("xlsx-formulas", "XLSX の数式", ("formulas",), "xlsx"),
    ("xlsx-objects", "XLSX の図形・画像等のオブジェクト", ("xlsx-object",), "xlsx"),
)


def _check_statuses(path: str, findings: list[Finding]) -> list[dict]:
    """Return every check outcome, including checks with no detected issue."""
    suffix = Path(path).suffix.lower()
    if suffix not in {".csv", ".tsv", ".xlsx", ".xls"} and ":" in path:
        suffix = Path(path.rsplit(":", 1)[0]).suffix.lower()
    has_table = suffix in {".csv", ".tsv", ".xlsx"}
    present_codes = {finding.code for finding in findings}
    statuses: list[dict] = []
    for check_id, label, codes, scope in CHECK_GROUPS:
        applicable = scope == "all" or scope == "table" and has_table or scope == "xlsx" and suffix == ".xlsx"
        detected = sorted(set(codes) & present_codes)
        statuses.append(
            {
                "id": check_id,
                "label": label,
                "status": "not_applicable" if not applicable else "issues_found" if detected else "passed",
                "finding_codes": detected,
                "finding_count": sum(finding.code in codes for finding in findings),
            }
        )
    return statuses


def _add(
    result: CheckResult,
    code: str,
    message: str,
    severity: str = "warning",
    row: int | None = None,
    column: int | None = None,
    value: str | None = None,
) -> None:
    result.findings.append(Finding(code, message, severity, row, column, value))


def check_rows(rows: Iterable[Iterable[object]], path: str = "<memory>") -> CheckResult:
    """Check a rectangular table; the first non-empty row is treated as its header."""
    data = [["" if cell is None else str(cell) for cell in row] for row in rows]
    result = CheckResult(path)
    if not data or not any(any(cell.strip() for cell in row) for row in data):
        _add(result, "empty-table", "表にデータがありません。", "error")
        return result

    first = next(i for i, row in enumerate(data) if any(cell.strip() for cell in row))
    if first:
        _add(result, "leading-empty-rows", "先頭の空白行は削除してください。", row=1)
    header = data[first]
    width = len(header)
    names: dict[str, int] = {}
    for col, name in enumerate(header, 1):
        clean = name.strip()
        if not clean:
            _add(result, "missing-header", "項目名を省略しないでください。", "error", first + 1, col, name)
        elif clean in names:
            _add(result, "duplicate-header", f"項目名「{clean}」が重複しています。", "error", first + 1, col, name)
        else:
            names[clean] = col

    gap_seen = False
    for index, row in enumerate(data[first + 1 :], first + 2):
        if not any(cell.strip() for cell in row):
            gap_seen = True
            continue
        if gap_seen:
            _add(result, "split-table", "空白行で表を分断しないでください。", "warning", index)
            gap_seen = False
        if len(row) != width:
            _add(result, "inconsistent-columns", "データ行の列数が項目名の列数と一致しません。", "error", index)
        for col, value in enumerate(row, 1):
            _check_cell(result, value, index, col)
    return result


def _check_cell(result: CheckResult, value: str, row: int, column: int) -> None:
    if not value:
        return
    if FORBIDDEN_LAYOUT.search(value):
        _add(result, "layout-whitespace", "空白や改行で体裁を整えず、列を分けてください。", "warning", row, column, value)
    if DEPENDENT_WORDS.search(value):
        _add(result, "dependent-character", "機種依存文字は使用しないでください。", "warning", row, column, value)
    if NUMERIC_WITH_DECORATION.match(value):
        _add(result, "decorated-number", "数値・単位・注記は別の列にしてください。", "warning", row, column, value)
    if ERA_ONLY.match(value.strip()):
        _add(result, "era-only-date", "時間軸は西暦を併記してください。", "warning", row, column, value)


def check_file(path: str | Path) -> CheckResult:
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        with file_path.open("r", encoding="utf-8-sig", newline="") as handle:
            return check_rows(csv.reader(handle, delimiter=delimiter), str(file_path))
    if suffix == ".xlsx":
        return _check_xlsx(file_path)
    result = CheckResult(str(file_path))
    if suffix == ".xls":
        _add(result, "legacy-xls", "古い .xls 形式は構造検査できません。.xlsx へ変換してください。", "warning")
    else:
        _add(result, "unsupported-format", "CSV、TSV、XLSX、XLS のいずれかを指定してください。", "error")
    return result


def _check_xlsx(path: Path) -> CheckResult:
    result = CheckResult(str(path))
    try:
        workbook = load_workbook(path, data_only=False, read_only=False)
        if not workbook.worksheets:
            _add(result, "empty-workbook", "ワークシートがありません。", "error")
            return result
        for sheet in workbook.worksheets:
            for merged_range in sheet.merged_cells.ranges:
                _add(
                    result,
                    "merged-cells",
                    "セル結合は使用しないでください。",
                    row=merged_range.min_row,
                    column=merged_range.min_col,
                    value=str(merged_range),
                )
            if sheet._images or sheet._charts:
                _add(result, "xlsx-object", "図形・画像等のオブジェクトではなくセルにデータを入力してください。")
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == "f":
                        _add(result, "formulas", "結果表は数式ではなく値として出力してください。", row=cell.row, column=cell.column, value=str(cell.value))
            rows = [[cell.value for cell in row] for row in sheet.iter_rows()]
            sheet_result = check_rows(rows, f"{path}:{sheet.title}")
            result.findings.extend(sheet_result.findings)
        workbook.close()
    except (OSError, InvalidFileException, ValueError) as error:
        _add(result, "invalid-xlsx", f"XLSX を読み取れません: {error}", "error")
    return result
